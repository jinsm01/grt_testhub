"""
批量生成「按模块导入点播课」模板测试数据。

读取原始导入模板 Excel，在保留表头/格式的前提下，
批量填充一级模块、二级模块字段，点播课名称和点播课ID留空供手动填写。

Usage:
    python generate_module_courses.py                          # 默认参数
    python generate_module_courses.py --primary-count 10 --secondary-count 5
    python generate_module_courses.py --output custom_output.xlsx
"""

import argparse
import sys
from typing import List, Tuple

import openpyxl


def generate_primary_module_names(count: int, max_length: int = 20) -> List[str]:
    """生成一级模块名称列表，保证不重复且不超过最大长度。

    Args:
        count: 需要生成的名称数量
        max_length: 每个名称的最大字符数

    Returns:
        不重复的一级模块名称列表
    """
    names: List[str] = []
    for i in range(1, count + 1):
        name = f"一级模块{i}测试数据验证"
        if len(name) > max_length:
            name = name[:max_length]
        names.append(name)
    return names


def generate_secondary_module_names(
    primary_names: List[str],
    count_per_primary: int,
    max_length: int = 20,
) -> List[Tuple[str, str]]:
    """为每个一级模块生成对应的二级模块名称，保证全局不重复。

    Args:
        primary_names: 一级模块名称列表
        count_per_primary: 每个一级模块下需要生成的二级模块数量
        max_length: 每个名称的最大字符数

    Returns:
        (一级模块名称, 二级模块名称) 的列表，二级模块全局唯一
    """
    rows: List[Tuple[str, str]] = []
    for p_idx, p_name in enumerate(primary_names, start=1):
        for s_idx in range(1, count_per_primary + 1):
            name = f"二级模块{p_idx}-{s_idx}专项学习课程"
            if len(name) > max_length:
                name = name[:max_length]
            rows.append((p_name, name))
    return rows


def fill_template(
    template_path: str,
    output_path: str,
    data: List[Tuple[str, str]],
    data_start_row: int = 3,
    primary_col: int = 1,
    secondary_col: int = 2,
) -> None:
    """将生成的数据填入模板并保存为新文件。

    Args:
        template_path: 原始模板 Excel 文件路径
        output_path: 输出文件路径
        data: (一级模块名称, 二级模块名称) 列表
        data_start_row: 数据起始行号（跳过表头和须知）
        primary_col: 一级模块所在列号
        secondary_col: 二级模块所在列号
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    for row_idx, (p_name, s_name) in enumerate(data, start=data_start_row):
        ws.cell(row=row_idx, column=primary_col, value=p_name)
        ws.cell(row=row_idx, column=secondary_col, value=s_name)

    wb.save(output_path)


def validate_data(
    data: List[Tuple[str, str]],
    primary_max_length: int = 20,
    secondary_max_length: int = 20,
) -> None:
    """校验生成的数据是否满足业务约束。

    Args:
        data: (一级模块名称, 二级模块名称) 列表
        primary_max_length: 一级模块名称最大长度
        secondary_max_length: 二级模块名称最大长度

    Raises:
        ValueError: 如果数据不满足约束条件
    """
    primary_names = [row[0] for row in data]
    secondary_names = [row[1] for row in data]

    for name in primary_names:
        if len(name) > primary_max_length:
            raise ValueError(f"一级模块名称超长: '{name}' ({len(name)}字)")

    for name in secondary_names:
        if len(name) > secondary_max_length:
            raise ValueError(f"二级模块名称超长: '{name}' ({len(name)}字)")

    if len(set(secondary_names)) != len(secondary_names):
        raise ValueError("二级模块名称存在重复")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="批量生成「按模块导入点播课」模板测试数据",
    )
    parser.add_argument(
        "--template",
        default="/Users/jinshaomin/Downloads/按模块导入点播课.xlsx",
        help="原始模板 Excel 文件路径",
    )
    parser.add_argument(
        "--output",
        default="/Users/jinshaomin/Downloads/按模块导入点播课_填充数据.xlsx",
        help="输出文件路径",
    )
    parser.add_argument(
        "--primary-count",
        type=int,
        default=50,
        help="一级模块数量（默认50）",
    )
    parser.add_argument(
        "--secondary-count",
        type=int,
        default=30,
        help="每个一级模块下的二级模块数量（默认30）",
    )
    parser.add_argument(
        "--primary-max-length",
        type=int,
        default=20,
        help="一级模块名称最大字数（默认20）",
    )
    parser.add_argument(
        "--secondary-max-length",
        type=int,
        default=20,
        help="二级模块名称最大字数（默认20）",
    )
    parser.add_argument(
        "--data-start-row",
        type=int,
        default=3,
        help="模板中数据起始行号（默认3）",
    )
    parser.add_argument(
        "--skip-validate",
        action="store_true",
        help="跳过数据校验",
    )
    args = parser.parse_args()

    total_rows = args.primary_count * args.secondary_count
    if total_rows > 2000:
        print(f"警告: 总数据量 {total_rows} 超过模板单次上限 2000 条，请调整参数")
        sys.exit(1)

    primary_names = generate_primary_module_names(args.primary_count, args.primary_max_length)
    data = generate_secondary_module_names(
        primary_names, args.secondary_count, args.secondary_max_length,
    )

    if not args.skip_validate:
        validate_data(data, args.primary_max_length, args.secondary_max_length)

    fill_template(
        args.template, args.output, data, args.data_start_row,
    )

    print(f"完成: {args.primary_count} 个一级模块 × {args.secondary_count} 个二级模块 = {total_rows} 条数据")
    print(f"输出文件: {args.output}")


if __name__ == "__main__":
    main()